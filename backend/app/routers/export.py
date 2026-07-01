"""
Router — export de dados em planilha Excel (.xlsx)

GET /api/export  →  retorna arquivo .xlsx com 4 abas:
  Ranking         — score_final, scores por categoria, total_matches
  Stats completas — todas as métricas por player (médias/somas)
  Histórico       — todas as partidas com stats individuais
  Head-to-Head    — matrix de kills entre todos os players

Acesso: público (os dados já são públicos via API)
"""

import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.services.ranking_service import get_ranking
from app.models.match import PlayerMatchStats, Match, PlayerVsPlayerStats
from app.models.player import Player

router = APIRouter(prefix="/api/export", tags=["export"])

# Paleta EverestFrags
_TEAL   = "FF0E7490"
_INDIGO = "FF6366F1"
_GOLD   = "FFE0A82E"
_DARK   = "FF0D1218"
_WHITE  = "FFF0F9FF"
_GRAY   = "FF1B2530"
_BG     = "FF070A0E"


def _header_style(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=_WHITE, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            bottom=Side(style="thin", color=_GRAY)
        )
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 12)


def _build_ranking_sheet(ws, entries):
    ws.title = "Ranking"
    headers = ["#", "Player", "Score Final", "Combate", "Duelos", "Utility", "Partidas", "K/D", "ADR", "Rating"]
    _header_style(ws, 1, headers)
    for i, e in enumerate(entries, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=e["rank"])
        ws.cell(row=row, column=2, value=e.get("player_display_name") or e["player_nickname"])
        ws.cell(row=row, column=3, value=round(e["score_final"], 1))
        ws.cell(row=row, column=4, value=round(e["score_combat"], 1))
        ws.cell(row=row, column=5, value=round(e["score_duel"], 1))
        ws.cell(row=row, column=6, value=round(e["score_utility"], 1))
        ws.cell(row=row, column=7, value=e["total_matches"])
        ws.cell(row=row, column=8, value=round(e["kd_ratio"], 2))
        ws.cell(row=row, column=9, value=round(e["adr"], 1))
        ws.cell(row=row, column=10, value=round(e["hltv_rating"], 2))
        # zebra
        if i % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FF0F1A23")


def _build_stats_sheet(ws, entries):
    ws.title = "Stats Completas"
    headers = [
        "Player", "Partidas", "Kills", "Deaths", "Assists", "K/D",
        "ADR", "Rating HLTV", "KAST%", "Opening K", "Opening D",
        "Trade K", "Trade Denial", "Flash Assists", "Grenade Dmg",
        "HE Acertos", "Fire Acertos", "Fire Dmg", "Eco Kills",
        "Disadv Kills", "Adv Kills", "MVPs", "TTK ms"
    ]
    _header_style(ws, 1, headers)
    for i, e in enumerate(entries, 1):
        r = i + 1
        vals = [
            e.get("player_display_name") or e["player_nickname"],
            e["total_matches"], e["kills"], e["deaths"], e["assists"],
            round(e["kd_ratio"], 2),
            round(e["adr"], 1), round(e["hltv_rating"], 2),
            round(e["kast_percent"], 1),
            e["opening_kills"], e.get("opening_deaths", 0),
            e["trade_kills"], e.get("trade_denials", 0),
            e["flash_assists"], e["grenade_damage"],
            e["he_enemies_hit"], e["fire_enemies_hit"], e.get("fire_damage", 0),
            e.get("eco_kills", 0), e.get("disadvantage_kills", 0),
            e.get("advantage_kills", 0), e.get("mvps", 0),
            round(e["time_to_kill_ms"]),
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=r, column=col, value=v)
        if i % 2 == 0:
            for col in range(1, len(vals) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FF0F1A23")


def _build_history_sheet(ws, db: Session):
    ws.title = "Histórico"
    headers = ["Data", "Mapa", "Player", "K", "D", "A", "ADR", "Rating", "KAST%", "Opening K", "Flash A", "MVPs"]
    _header_style(ws, 1, headers)

    rows = (
        db.query(PlayerMatchStats, Match, Player)
        .join(Match, PlayerMatchStats.match_id == Match.id)
        .join(Player, PlayerMatchStats.player_id == Player.id)
        .filter(Player.is_active == True)  # noqa: E712
        .order_by(Match.played_at.desc(), Match.id.desc(), Player.nickname)
        .all()
    )

    for i, (stat, match, player) in enumerate(rows, 1):
        r = i + 1
        vals = [
            str(match.played_at),
            match.map_name or "—",
            player.display_name or player.nickname,
            stat.kills, stat.deaths, stat.assists,
            round(float(stat.adr), 1),
            round(float(stat.hltv_rating), 2),
            round(float(stat.kast_percent), 1),
            stat.opening_kills,
            stat.flash_assists,
            stat.mvps,
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=r, column=col, value=v)
        if i % 2 == 0:
            for col in range(1, len(vals) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FF0F1A23")


def _build_h2h_sheet(ws, db: Session, entries):
    ws.title = "Head-to-Head"
    players = [
        {"id": e["player_id"], "name": e.get("player_display_name") or e["player_nickname"]}
        for e in entries
    ]
    n = len(players)

    # Cabeçalho de coluna
    ws.cell(row=1, column=1, value="↓ matou →").font = Font(bold=True, color=_WHITE)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=_DARK)
    for col, p in enumerate(players, 2):
        cell = ws.cell(row=1, column=col, value=p["name"])
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 14

    for row, p in enumerate(players, 2):
        cell = ws.cell(row=row, column=1, value=p["name"])
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_TEAL)

    # Preenche kills de cada par
    all_vs = db.query(PlayerVsPlayerStats).all()
    kills_map: dict[tuple, int] = {}
    for vs in all_vs:
        kills_map[(vs.player_id, vs.opponent_id)] = kills_map.get((vs.player_id, vs.opponent_id), 0) + vs.kills

    for r, attacker in enumerate(players, 2):
        for c, victim in enumerate(players, 2):
            if attacker["id"] == victim["id"]:
                cell = ws.cell(row=r, column=c, value="—")
                cell.fill = PatternFill("solid", fgColor=_DARK)
            else:
                k = kills_map.get((attacker["id"], victim["id"]), 0)
                ws.cell(row=r, column=c, value=k if k else "")

    ws.column_dimensions["A"].width = 18


def generate_xlsx(db: Session) -> bytes:
    entries = get_ranking(db)
    # Converte RankingEntry para dict
    if entries and hasattr(entries[0], "model_dump"):
        entries = [e.model_dump() for e in entries]

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    ws_rank = wb.create_sheet()
    _build_ranking_sheet(ws_rank, entries)

    ws_stats = wb.create_sheet()
    _build_stats_sheet(ws_stats, entries)

    ws_hist = wb.create_sheet()
    _build_history_sheet(ws_hist, db)

    ws_h2h = wb.create_sheet()
    _build_h2h_sheet(ws_h2h, db, entries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("")
def export_xlsx(db: Session = Depends(get_db)):
    """Exporta todos os dados do grupo em planilha .xlsx (4 abas)."""
    data = generate_xlsx(db)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=everestfrags.xlsx"},
    )
